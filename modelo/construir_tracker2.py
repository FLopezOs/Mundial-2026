# -*- coding: utf-8 -*-
"""Parte 2: hojas Modelo, Predicciones y Matrices (fórmulas vivas)."""
import os, sys
sys.path.insert(0, os.path.dirname(__file__))
from comun import cargar_fixture
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.join(os.path.dirname(__file__), "..")
F_HDR = Font(name="Arial", size=10, bold=True, color="FFFFFF")
F_STD = Font(name="Arial", size=10)
F_NOTA = Font(name="Arial", size=9, italic=True, color="808080")
FILL_HDR = PatternFill("solid", start_color="1F4E78")
FILL_GRID = PatternFill("solid", start_color="F2F2F2")
CENTRO = Alignment(horizontal="center")

wb = load_workbook(os.path.join(BASE, "output", "Mundial2026_Tracker.xlsx"))

# ============ Hoja Modelo ============
# fila 1: encabezados | fila 2: INICIAL (pre-torneo) | filas 3..106: partidos 1..104
ws = wb.create_sheet("Modelo")
cols = ["ID","Fecha","Fase","Equipo_A","Equipo_B","Loc_A","Loc_B","Jugado","GA","GB",
        "Elo_A_pre","Elo_B_pre","dr","E_A","S_A","G_mult","Delta_A","Elo_A_post","Elo_B_post",
        "Δ_goles","Att_A","Def_A","Att_B","Def_B","Total","λ_A","λ_B"]
for j, t in enumerate(cols, 1):
    c = ws.cell(row=1, column=j, value=t)
    c.font, c.fill, c.alignment = F_HDR, FILL_HDR, CENTRO
ws.cell(row=2, column=1, value="INICIAL").font = Font(name="Arial", size=10, bold=True)

# grid de cadena Elo: AE..BZ (48 equipos), fila 2 = Elo_Inicial, fila r = Elo tras partido r-2
for j in range(31, 79):  # AE=31 .. BZ=78
    L = get_column_letter(j)
    c = ws.cell(row=1, column=j, value=f"=Equipos!$A{j-29}")  # equipos en Equipos!A2:A49
    c.font, c.fill = F_HDR, FILL_HDR
    ws.cell(row=2, column=j, value=f"=VLOOKUP({L}$1,Equipos!$A:$C,3,0)").number_format = "0.0"

for i in range(1, 105):
    r = i + 2
    fr = i + 1   # fila en Fixture/Resultados
    ws.cell(row=r, column=1, value=f"=Fixture!A{fr}")
    ws.cell(row=r, column=2, value=f"=Fixture!B{fr}")
    ws.cell(row=r, column=3, value=f"=Fixture!C{fr}")
    ws.cell(row=r, column=4, value=f"=Fixture!E{fr}")
    ws.cell(row=r, column=5, value=f"=Fixture!F{fr}")
    ws.cell(row=r, column=6, value=f"=Fixture!J{fr}")
    ws.cell(row=r, column=7, value=f"=Fixture!K{fr}")
    ws.cell(row=r, column=8, value=f"=IF(ISNUMBER(Resultados!B{fr}),1,0)")
    ws.cell(row=r, column=9, value=f"=IF($H{r}=1,Resultados!B{fr},0)")
    ws.cell(row=r, column=10, value=f"=IF($H{r}=1,Resultados!C{fr},0)")
    ws.cell(row=r, column=11, value=f"=IFERROR(INDEX($AE{r-1}:$BZ{r-1},MATCH($D{r},$AE$1:$BZ$1,0)),\"\")")
    ws.cell(row=r, column=12, value=f"=IFERROR(INDEX($AE{r-1}:$BZ{r-1},MATCH($E{r},$AE$1:$BZ$1,0)),\"\")")
    ws.cell(row=r, column=13, value=f"=IF(OR($K{r}=\"\",$L{r}=\"\"),\"\",($K{r}+100*$F{r})-($L{r}+100*$G{r}))")
    ws.cell(row=r, column=14, value=f"=IF($M{r}=\"\",\"\",1/(1+10^(-$M{r}/400)))")
    ws.cell(row=r, column=15, value=f"=IF(OR($H{r}=0,$N{r}=\"\"),\"\",IF($I{r}>$J{r},1,IF($I{r}=$J{r},0.5,0)))")
    ws.cell(row=r, column=16, value=f"=IF($O{r}=\"\",\"\",IF(ABS($I{r}-$J{r})<=1,1,IF(ABS($I{r}-$J{r})=2,1.5,IF(ABS($I{r}-$J{r})=3,1.75,1.75+(ABS($I{r}-$J{r})-3)/8))))")
    ws.cell(row=r, column=17, value=f"=IF($O{r}=\"\",0,Equipos!$O$4*$P{r}*($O{r}-$N{r}))")
    ws.cell(row=r, column=18, value=f"=IF($K{r}=\"\",\"\",$K{r}+$Q{r})")
    ws.cell(row=r, column=19, value=f"=IF($L{r}=\"\",\"\",$L{r}-$Q{r})")
    ws.cell(row=r, column=20, value=f"=IF($M{r}=\"\",\"\",Equipos!$O$3*$M{r})")
    ws.cell(row=r, column=21, value=f"=IFERROR(VLOOKUP($D{r},Equipos!$A:$H,6,0),\"\")")
    ws.cell(row=r, column=22, value=f"=IFERROR(VLOOKUP($D{r},Equipos!$A:$H,7,0),\"\")")
    ws.cell(row=r, column=23, value=f"=IFERROR(VLOOKUP($E{r},Equipos!$A:$H,6,0),\"\")")
    ws.cell(row=r, column=24, value=f"=IFERROR(VLOOKUP($E{r},Equipos!$A:$H,7,0),\"\")")
    ws.cell(row=r, column=25, value=f"=IF(OR($U{r}=\"\",$W{r}=\"\"),\"\",MIN(3.6,MAX(1.8,Equipos!$O$2*SQRT($U{r}*$W{r}*$V{r}*$X{r}))))")
    ws.cell(row=r, column=26, value=f"=IF(OR($Y{r}=\"\",$T{r}=\"\"),\"\",MAX(0.2,($Y{r}+$T{r})/2))")
    ws.cell(row=r, column=27, value=f"=IF(OR($Y{r}=\"\",$T{r}=\"\"),\"\",MAX(0.2,($Y{r}-$T{r})/2))")
    # grid
    for j in range(31, 79):
        L = get_column_letter(j)
        gc = ws.cell(row=r, column=j, value=f"=IF($D{r}={L}$1,$R{r},IF($E{r}={L}$1,$S{r},{L}{r-1}))")
        gc.number_format = "0.0"
        gc.fill = FILL_GRID
    for j, fmt in ((11,"0.0"),(12,"0.0"),(13,"0.0"),(14,"0.0%"),(16,"0.00"),(17,"0.00"),
                   (18,"0.0"),(19,"0.0"),(20,"0.00"),(21,"0.00"),(22,"0.00"),(23,"0.00"),
                   (24,"0.00"),(25,"0.00"),(26,"0.00"),(27,"0.00")):
        ws.cell(row=r, column=j).number_format = fmt
for r in range(2, 107):
    for j in range(1, 28):
        if ws.cell(row=r, column=j).font.name != "Arial":
            ws.cell(row=r, column=j).font = F_STD
ws.cell(row=108, column=1, value="Solo lectura conceptual: nadie edita aquí. Grid AE:BZ = Elo encadenado (fila 2 = inicial; fila r = Elo tras el partido de la fila r). Elo pre-partido = fila anterior del grid.").font = F_NOTA
for col, w in zip(["A","B","C","D","E"], [6,11,9,18,18]):
    ws.column_dimensions[col].width = w
for j in range(6, 28):
    ws.column_dimensions[get_column_letter(j)].width = 8.5
for j in range(31, 79):
    ws.column_dimensions[get_column_letter(j)].width = 9
ws.freeze_panes = "F3"

# ============ Hoja Predicciones ============
ws = wb.create_sheet("Predicciones")
ws.cell(row=1, column=1, value="NOTA: probabilidades de un modelo Elo+Poisson calibrado con datos históricos; no incorpora lesiones, alineaciones ni clima, y en fútbol el favorito pierde con frecuencia. P(1X2) normalizadas sobre marcadores 0–5 (residuo >5 repartido proporcionalmente). Marcador más probable = moda de las Poisson.").font = F_NOTA
cols = ["ID","Fecha","Fase","Equipo_A","Equipo_B","P(Gana A)","P(Empate)","P(Gana B)",
        "Marcador_Mas_Probable","P(Marcador)","raw_pA (aux)","raw_pE (aux)","raw_pB (aux)"]
for j, t in enumerate(cols, 1):
    c = ws.cell(row=2, column=j, value=t)
    c.font, c.fill, c.alignment = F_HDR, FILL_HDR, CENTRO
for i in range(1, 105):
    r = i + 2
    mr = i + 2  # fila en Modelo
    cond = f"OR(Modelo!$H{mr}=1,Modelo!$Z{mr}=\"\")"
    ws.cell(row=r, column=1, value=f"=Modelo!A{mr}")
    ws.cell(row=r, column=2, value=f"=Modelo!B{mr}")
    ws.cell(row=r, column=3, value=f"=Modelo!C{mr}")
    ws.cell(row=r, column=4, value=f"=Modelo!D{mr}")
    ws.cell(row=r, column=5, value=f"=Modelo!E{mr}")
    ws.cell(row=r, column=11, value=f"=IF({cond},\"\",SUMPRODUCT(POISSON({{1,2,3,4,5}},Modelo!$Z{mr},0),POISSON({{0,1,2,3,4}},Modelo!$AA{mr},1)))")
    ws.cell(row=r, column=12, value=f"=IF({cond},\"\",SUMPRODUCT(POISSON({{0,1,2,3,4,5}},Modelo!$Z{mr},0),POISSON({{0,1,2,3,4,5}},Modelo!$AA{mr},0)))")
    ws.cell(row=r, column=13, value=f"=IF({cond},\"\",SUMPRODUCT(POISSON({{1,2,3,4,5}},Modelo!$AA{mr},0),POISSON({{0,1,2,3,4}},Modelo!$Z{mr},1)))")
    ws.cell(row=r, column=6, value=f"=IF($K{r}=\"\",\"\",$K{r}/($K{r}+$L{r}+$M{r}))")
    ws.cell(row=r, column=7, value=f"=IF($K{r}=\"\",\"\",$L{r}/($K{r}+$L{r}+$M{r}))")
    ws.cell(row=r, column=8, value=f"=IF($K{r}=\"\",\"\",$M{r}/($K{r}+$L{r}+$M{r}))")
    ws.cell(row=r, column=9, value=f"=IF($K{r}=\"\",\"\",INT(Modelo!$Z{mr})&\"-\"&INT(Modelo!$AA{mr}))")
    ws.cell(row=r, column=10, value=f"=IF($K{r}=\"\",\"\",POISSON(INT(Modelo!$Z{mr}),Modelo!$Z{mr},0)*POISSON(INT(Modelo!$AA{mr}),Modelo!$AA{mr},0))")
    for j in (6,7,8,10,11,12,13):
        ws.cell(row=r, column=j).number_format = "0.0%"
    for j in range(1, 14):
        ws.cell(row=r, column=j).font = F_STD
for col, w in zip("ABCDEFGHIJKLM", [6,11,9,18,18,11,11,11,12,11,10,10,10]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A3"

# ============ Hoja Matrices ============
ws = wb.create_sheet("Matrices")
ws.cell(row=1, column=1, value="MATRIZ DE MARCADORES 0–5 × 0–5").font = Font(name="Arial", size=11, bold=True)
ws.cell(row=2, column=1, value="ID_Partido:").font = F_STD
sel = ws.cell(row=2, column=2, value=4)
sel.fill = PatternFill("solid", start_color="FFF2CC"); sel.font = F_STD
ws.cell(row=2, column=3, value='=VLOOKUP($B$2,Fixture!$A:$F,5,0)&" vs "&VLOOKUP($B$2,Fixture!$A:$F,6,0)').font = F_STD
ws.cell(row=3, column=1, value="λ_A:").font = F_STD
ws.cell(row=3, column=2, value="=INDEX(Modelo!$Z$3:$Z$106,MATCH($B$2,Modelo!$A$3:$A$106,0))").number_format = "0.00"
ws.cell(row=3, column=3, value="λ_B:").font = F_STD
ws.cell(row=3, column=4, value="=INDEX(Modelo!$AA$3:$AA$106,MATCH($B$2,Modelo!$A$3:$A$106,0))").number_format = "0.00"
ws.cell(row=5, column=1, value="A \\ B").font = F_HDR; ws.cell(row=5, column=1).fill = FILL_HDR
for k in range(6):
    c = ws.cell(row=5, column=2+k, value=k); c.font, c.fill, c.alignment = F_HDR, FILL_HDR, CENTRO
    c2 = ws.cell(row=6+k, column=1, value=k); c2.font, c2.fill, c2.alignment = F_HDR, FILL_HDR, CENTRO
for a in range(6):
    for b in range(6):
        c = ws.cell(row=6+a, column=2+b, value=f"=POISSON({a},$B$3,0)*POISSON({b},$D$3,0)")
        c.number_format = "0.0%"; c.font = F_STD
ws.cell(row=13, column=1, value="Residuo (>5 goles): 1 − suma de la matriz =").font = F_NOTA
ws.cell(row=13, column=4, value="=1-SUM(B6:G11)").number_format = "0.0%"
for col, w in zip("ABCDEFG", [10,8,8,8,8,8,8]):
    ws.column_dimensions[col].width = w

wb.save(os.path.join(BASE, "output", "Mundial2026_Tracker.xlsx"))
print("Parte 2 OK: Modelo, Predicciones, Matrices")
