# -*- coding: utf-8 -*-
"""
actualizar.py — Rutina post-jornada (INSTRUCCIONES.md §7).
Fuente principal: ESPN unofficial API (cerca de tiempo real).
Fallback:         openfootball worldcup.json (actualización manual diaria).

Flujo:
1) Intenta ESPN API para todas las fechas desde inicio del torneo hasta hoy.
2) Si ESPN falla totalmente, cae a openfootball worldcup.json.
3) Carga SOLO partidos nuevos en Resultados (nunca pisa digitado a mano; conflicto → reporta).
4) Actualiza nombres en Fixture cuando el json resuelve placeholders de eliminatorias.
5) Verifica recálculo Elo con cadena independiente en Python y reporta movimientos.

NOTA: ESPN scoreboard a veces no actualiza status.type.completed aunque el partido
terminó (bug conocido). es_finalizado() usa múltiples señales para detectar FT.
"""
import hashlib, json, os, shutil, sys, urllib.request
from datetime import datetime, timedelta, timezone, date as date_t
sys.path.insert(0, os.path.dirname(__file__))
from comun import cargar_fixture, canon, JSON_A_CANON
from openpyxl import load_workbook
try:
    sys.stdout.reconfigure(encoding="utf-8")
except AttributeError:
    pass

BASE = os.path.join(os.path.dirname(__file__), "..")
DATA = os.path.join(BASE, "Data")
XLSX = os.path.join(BASE, "output", "Mundial2026_Tracker.xlsx")

# ── ESPN API ────────────────────────────────────────────────────────────────
ESPN_URL = "https://site.api.espn.com/apis/site/v2/sports/soccer/fifa.world/scoreboard?dates={}"
TORNEO_INICIO = date_t(2026, 6, 11)

# ESPN displayName → canónico ES  (extiende JSON_A_CANON para variantes ESPN)
ESPN_EXTRA = {
    "United States":         "Estados Unidos",
    "Ivory Coast":           "Costa de Marfil",
    "DR Congo":              "RD Congo",
    "Congo DR":              "RD Congo",
    "Democratic Republic of Congo": "RD Congo",
    "South Korea":           "Corea del Sur",
    "Haiti":                 "Haití",
    "Bosnia-Herzegovina":    "Bosnia y Herzegovina",
    "Bosnia and Herzegovina":"Bosnia y Herzegovina",
    "Czechia":               "Chequia",
    "Czech Republic":        "Chequia",
    "Iran":                  "Irán",
    "Iraq":                  "Irak",
    "Japan":                 "Japón",
    "Jordan":                "Jordania",
    "Mexico":                "México",
    "Morocco":               "Marruecos",
    "Netherlands":           "Países Bajos",
    "New Zealand":           "Nueva Zelanda",
    "Norway":                "Noruega",
    "Panama":                "Panamá",
    "Saudi Arabia":          "Arabia Saudita",
    "Scotland":              "Escocia",
    "South Africa":          "Sudáfrica",
    "Switzerland":           "Suiza",
    "Tunisia":               "Túnez",
    "Turkey":                "Turquía",
    "Türkiye":               "Turquía",
    "Uzbekistan":            "Uzbekistán",
    "Algeria":               "Argelia",
    "Cape Verde":            "Cabo Verde",
}

def espn_canon(nombre):
    """ESPN displayName → canónico ES."""
    if nombre in ESPN_EXTRA:
        return ESPN_EXTRA[nombre]
    return JSON_A_CANON.get(nombre, nombre)

# Nombres de estado ESPN que indican partido terminado
_ESTADOS_FINALES = {
    "STATUS_FULL_TIME", "STATUS_FINAL", "STATUS_FINAL_EXTRA_TIME",
    "STATUS_FINAL_AET", "STATUS_FINAL_OVERTIME", "STATUS_AP",
    "STATUS_FT", "STATUS_SHOOTOUT",
}
# Nombres que explícitamente indican partido NO terminado (vivo o por jugar)
_ESTADOS_EN_CURSO = {
    "STATUS_SCHEDULED", "STATUS_IN_PROGRESS",
    "STATUS_HALFTIME", "STATUS_DELAYED", "STATUS_POSTPONED",
    "STATUS_CANCELED",
}

def es_finalizado(ev, comp):
    """
    Determina si un partido terminó.
    ESPN a veces devuelve completed=False aunque el partido ya terminó.
    Usa 3 señales en cascada para ser robusto.
    """
    tipo = comp["status"]["type"]

    # Señal 1: ESPN lo marca explícitamente
    if tipo.get("completed", False):
        return True

    estado = tipo.get("name", "")

    # Señal 2: nombre de estado conocido como "terminado"
    if estado in _ESTADOS_FINALES:
        return True
    # Nombre contiene palabras clave de final (cobertura de variantes ESPN)
    if any(k in estado for k in ("FULL_TIME", "FINAL", "_AET", "SHOOTOUT")):
        return True

    # Señal 3: estado explícito de "no terminado" → salir
    if estado in _ESTADOS_EN_CURSO:
        return False

    # Señal 4 (fallback temporal): el evento fue hace >3 h y hay scores registrados
    try:
        ev_dt = datetime.fromisoformat(ev["date"].replace("Z", "+00:00"))
        elapsed = datetime.now(tz=timezone.utc) - ev_dt
        if elapsed.total_seconds() > 10_800:  # 3 horas
            scores = [c.get("score") for c in comp.get("competitors", [])]
            if all(s is not None for s in scores):
                return True
    except Exception:
        pass

    return False


def detectar_tipo(comp):
    """Detecta si el resultado fue en 90', ET o penales."""
    estado = comp["status"]["type"]["name"]
    # ESPN usa STATUS_FULL_TIME para FT normal
    # Para ET y penales hay otros estados; mapeamos lo conocido
    if "SHOOTOUT" in estado or "PENALTY" in estado:
        return "PEN"
    if "EXTRA_TIME" in estado or "OVERTIME" in estado:
        return "ET"
    return "90"

def descargar_espn():
    """
    Descarga resultados ESPN para todas las fechas del torneo hasta hoy.
    Retorna lista de dicts: {fecha, home, away, gHome, gAway, tipo}.
    fecha = YYYY-MM-DD en UTC (puede diferir 1 día del horario local CL/US).
    """
    today = date_t.today()
    current = TORNEO_INICIO
    partidos = []
    fechas_ok = 0

    while current <= today:
        fecha_str = current.strftime("%Y%m%d")
        url = ESPN_URL.format(fecha_str)
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            resp = urllib.request.urlopen(req, timeout=20).read()
            data = json.loads(resp)
            for ev in data.get("events", []):
                comp = ev["competitions"][0]
                if not es_finalizado(ev, comp):
                    continue
                teams = {}
                for c in comp["competitors"]:
                    teams[c["homeAway"]] = {
                        "name":  espn_canon(c["team"]["displayName"]),
                        "score": int(float(c.get("score", 0) or 0)),
                    }
                home = teams.get("home")
                away = teams.get("away")
                if not home or not away:
                    continue
                partidos.append({
                    "fecha": ev["date"][:10],   # YYYY-MM-DD UTC
                    "home":  home["name"],
                    "away":  away["name"],
                    "gHome": home["score"],
                    "gAway": away["score"],
                    "tipo":  detectar_tipo(comp),
                })
            fechas_ok += 1
        except Exception as e:
            print(f"[AVISO] ESPN {fecha_str}: {e}")
        current += timedelta(days=1)

    if fechas_ok == 0:
        return None   # señal de fallo total
    return partidos

# ── openfootball (fallback) ──────────────────────────────────────────────────
URL_OPENFOOTBALL = "https://raw.githubusercontent.com/openfootball/worldcup.json/master/2026/worldcup.json"

def descargar_openfootball():
    try:
        req = urllib.request.Request(URL_OPENFOOTBALL, headers={"User-Agent": "Mozilla/5.0"})
        data = urllib.request.urlopen(req, timeout=30).read()
        json.loads(data)
        with open(os.path.join(DATA, "worldcup.json"), "wb") as f:
            f.write(data)
        print("[OK] worldcup.json (openfootball) descargado como fallback.")
        return True
    except Exception as e:
        print(f"[AVISO] openfootball también falló ({e}). Usando local.")
        return False

# ── backup ───────────────────────────────────────────────────────────────────
def _md5(path):
    h = hashlib.md5()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()

def backup():
    bdir = os.path.join(BASE, "output", "backups")
    os.makedirs(bdir, exist_ok=True)
    bks = sorted(f for f in os.listdir(bdir) if f.startswith("Tracker_") and f.endswith(".xlsx"))
    if bks and _md5(XLSX) == _md5(os.path.join(bdir, bks[-1])):
        print("[OK] Backup omitido: sin cambios desde el último.")
        return
    dst = os.path.join(bdir, f"Tracker_{datetime.now():%Y%m%d_%H%M}.xlsx")
    shutil.copy2(XLSX, dst)
    bks = sorted(f for f in os.listdir(bdir) if f.startswith("Tracker_") and f.endswith(".xlsx"))
    for old in bks[:-10]:
        os.remove(os.path.join(bdir, old))
    print(f"[OK] Backup: {os.path.basename(dst)}")

# ── escritura en xlsx ─────────────────────────────────────────────────────────
def escribir_resultados(wb, partidos_espn=None, fx_fallback=None):
    """
    Escribe resultados nuevos en Resultados y actualiza Fixture.
    partidos_espn: lista de dicts ESPN (fuente principal)
    fx_fallback:   lista de cargar_fixture() (openfootball)
    """
    wsF, wsR = wb["Fixture"], wb["Resultados"]

    # Índice por (fecha, equipoA, equipoB) → fila
    fila_de = {}
    for r in range(2, 106):
        f  = wsF.cell(r, 2).value
        eA = wsF.cell(r, 5).value
        eB = wsF.cell(r, 6).value
        if f and eA and eB:
            fila_de[(f, eA, eB)] = r

    nuevos, conflictos, renombrados = [], [], 0

    if partidos_espn is not None:
        # ── FUENTE: ESPN ─────────────────────────────────────────────────────
        for p in partidos_espn:
            r = None
            gA = gB = None
            home, away = p["home"], p["away"]
            fecha = p["fecha"]
            # ESPN fecha UTC puede estar 1 día adelante respecto al Fixture (horario local)
            fechas = [fecha]
            dt = datetime.strptime(fecha, "%Y-%m-%d")
            fechas.append((dt - timedelta(days=1)).strftime("%Y-%m-%d"))

            for f in fechas:
                if (f, home, away) in fila_de:
                    r = fila_de[(f, home, away)]
                    gA, gB = p["gHome"], p["gAway"]
                    break
                if (f, away, home) in fila_de:
                    r = fila_de[(f, away, home)]
                    gA, gB = p["gAway"], p["gHome"]
                    break

            if r is None:
                print(f"[AVISO] Sin fila en Fixture: {fecha} {home} vs {away}")
                continue

            existA = wsR.cell(r, 2).value
            existB = wsR.cell(r, 3).value
            if existA is None and existB is None:
                wsR.cell(r, 2).value = gA
                wsR.cell(r, 3).value = gB
                wsR.cell(r, 4).value = p["tipo"]
                nuevos.append((wsF.cell(r, 1).value,
                               wsF.cell(r, 5).value, gA, gB, wsF.cell(r, 6).value))
            elif (existA, existB) != (gA, gB):
                conflictos.append((wsF.cell(r, 1).value,
                                   wsF.cell(r, 5).value, wsF.cell(r, 6).value,
                                   (existA, existB), (gA, gB)))
    else:
        # ── FUENTE: openfootball ──────────────────────────────────────────────
        for p in fx_fallback:
            key = (p["fecha"], p["a"], p["b"])
            r = fila_de.get(key)
            if r is None and p["num"]:
                for rr in range(2, 106):
                    if wsF.cell(rr, 12).value == p["num"]:
                        if (wsF.cell(rr, 5).value != p["a"] or
                                wsF.cell(rr, 6).value != p["b"]):
                            wsF.cell(rr, 5).value = p["a"]
                            wsF.cell(rr, 6).value = p["b"]
                            renombrados += 1
                        r = rr
                        break
            if r is None or not p["score"]:
                continue
            sc = p["score"]
            goles = sc.get("et") or sc["ft"]
            gA, gB = wsR.cell(r, 2).value, wsR.cell(r, 3).value
            if gA is None and gB is None:
                wsR.cell(r, 2).value = int(goles[0])
                wsR.cell(r, 3).value = int(goles[1])
                if sc.get("pen"):
                    wsR.cell(r, 4).value = "PEN"
                    pen = sc["pen"]
                    wsR.cell(r, 5).value = p["a"] if pen[0] > pen[1] else p["b"]
                elif sc.get("et"):
                    wsR.cell(r, 4).value = "ET"
                else:
                    wsR.cell(r, 4).value = "90"
                nuevos.append((wsF.cell(r, 1).value, p["a"], goles[0], goles[1], p["b"]))
            elif (gA, gB) != (int(goles[0]), int(goles[1])):
                conflictos.append((wsF.cell(r, 1).value, p["a"], p["b"],
                                   (gA, gB), tuple(goles)))
        if renombrados:
            print(f"Placeholders de eliminatorias resueltos en Fixture: {renombrados}")

    return nuevos, conflictos

# ── verificación Elo ──────────────────────────────────────────────────────────
def verificar_elo(wb):
    wsF, wsR, wsE = wb["Fixture"], wb["Resultados"], wb["Equipos"]
    elo = {wsE.cell(r, 1).value: float(wsE.cell(r, 3).value) for r in range(2, 50)
           if wsE.cell(r, 1).value}
    elo_ini = dict(elo)
    K = 60
    for r in range(2, 106):
        a, b = wsF.cell(r, 5).value, wsF.cell(r, 6).value
        ga, gb = wsR.cell(r, 2).value, wsR.cell(r, 3).value
        if a not in elo or b not in elo or ga is None:
            continue
        la = 1 if (a == wsF.cell(r, 9).value) else 0
        lb = 1 if (b == wsF.cell(r, 9).value) else 0
        dr = (elo[a] + 100 * la) - (elo[b] + 100 * lb)
        ea = 1 / (1 + 10 ** (-dr / 400))
        s  = 1.0 if ga > gb else (0.5 if ga == gb else 0.0)
        d  = abs(ga - gb)
        g  = 1 if d <= 1 else (1.5 if d == 2 else (1.75 if d == 3 else 1.75 + (d-3)/8))
        delta = K * g * (s - ea)
        elo[a] += delta
        elo[b] -= delta
    return elo, elo_ini

# ── main ──────────────────────────────────────────────────────────────────────
def main():
    print("=== Actualizando resultados ===")

    # 1. Intentar ESPN
    print("Consultando ESPN API...")
    partidos_espn = descargar_espn()

    fx_fallback = None
    fuente = "ESPN"
    if partidos_espn is None:
        print("[AVISO] ESPN no disponible. Usando openfootball como fallback.")
        descargar_openfootball()
        fx_fallback = cargar_fixture(os.path.join(DATA, "worldcup.json"))
        fuente = "openfootball"
    else:
        completados = len(partidos_espn)
        print(f"[OK] ESPN: {completados} partido(s) terminado(s) encontrado(s).")

    # 2. Backup + abrir xlsx
    backup()
    wb = load_workbook(XLSX)

    # 3. Escribir resultados
    nuevos, conflictos = escribir_resultados(
        wb,
        partidos_espn=partidos_espn,
        fx_fallback=fx_fallback
    )

    # 4. Cuotas de casas de apuesta (antes de guardar para ver partidos pendientes correctos)
    if partidos_espn is not None:
        print("Descargando cuotas (DraftKings vía ESPN)...")
        guardar_odds(wb)

    wb.save(XLSX)

    # 4. Verificación Elo (abre el libro recién guardado)
    wb2 = load_workbook(XLSX)
    elo, elo_ini = verificar_elo(wb2)

    # 5. Reporte
    print(f"\n--- REPORTE (fuente: {fuente}) ---")
    print(f"Resultados nuevos cargados: {len(nuevos)}")
    for n in nuevos:
        print(f"  ID {n[0]}: {n[1]} {n[2]}-{n[3]} {n[4]}")
    if conflictos:
        print("CONFLICTOS (valor manual no pisado; revisar con el usuario):")
        for c in conflictos:
            print(f"  ID {c[0]}: {c[1]} vs {c[2]} | libro={c[3]} ESPN={c[4]}")
    movs = sorted(elo, key=lambda t: -abs(elo[t] - elo_ini[t]))[:5]
    print("Top movimientos Elo (cadena Python; comparar contra Equipos!D tras recalcular):")
    for t in movs:
        if abs(elo[t] - elo_ini[t]) > 0.01:
            print(f"  {t}: {elo_ini[t]:.1f} -> {elo[t]:.1f} ({elo[t]-elo_ini[t]:+.1f})")
    if not nuevos and not conflictos:
        print("Sin partidos nuevos — todo ya estaba cargado o no hay resultados recientes.")
    print("\nListo! Recarga el browser (F5) para ver probabilities actualizadas.")
    print("(Opcional: recalcular el xlsx en Excel para refrescar fórmulas Elo en Equipos!D)")

# ── ODDS (casas de apuesta) ───────────────────────────────────────────────────
TORNEO_FIN = date_t(2026, 7, 19)

def _ml_to_prob(ml):
    """Moneyline americano (int o str) → probabilidad implícita cruda."""
    v = int(str(ml).replace("+", ""))
    return (100 / (v + 100)) if v > 0 else (abs(v) / (abs(v) + 100))

def guardar_odds(wb):
    """
    Descarga cuotas DraftKings vía ESPN para partidos pendientes y guarda Data/odds.json.
    Clave en el JSON: str(match_id) del Fixture.
    """
    wsF, wsR = wb["Fixture"], wb["Resultados"]

    # Índice solo de partidos SIN resultado
    fila_de_pend = {}
    for r in range(2, 106):
        if wsR.cell(r, 2).value is not None:
            continue  # ya tiene resultado
        f   = wsF.cell(r, 2).value
        eA  = wsF.cell(r, 5).value
        eB  = wsF.cell(r, 6).value
        pid = wsF.cell(r, 1).value
        if f and eA and eB:
            fila_de_pend[(str(f), eA, eB)] = pid

    today   = date_t.today()
    current = today
    odds_result = {}

    while current <= TORNEO_FIN:
        fecha_str = current.strftime("%Y%m%d")
        url = ESPN_URL.format(fecha_str)
        try:
            req  = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            resp = urllib.request.urlopen(req, timeout=15).read()
            data = json.loads(resp)

            for ev in data.get("events", []):
                comp      = ev["competitions"][0]
                odds_list = comp.get("odds", [])
                if not odds_list or not odds_list[0]:
                    continue

                odds     = odds_list[0]
                ml_obj   = odds.get("moneyline", {})
                draw_obj = odds.get("drawOdds", {})

                ml_home = ml_obj.get("home", {}).get("close", {}).get("odds")
                ml_away = ml_obj.get("away", {}).get("close", {}).get("odds")
                ml_draw = draw_obj.get("moneyLine")

                if not all([ml_home, ml_away, ml_draw is not None]):
                    continue

                try:
                    pH  = _ml_to_prob(ml_home)
                    pA  = _ml_to_prob(ml_away)
                    pD  = _ml_to_prob(ml_draw)
                    tot = pH + pA + pD
                    pH, pA, pD = round(pH/tot, 4), round(pA/tot, 4), round(pD/tot, 4)
                except Exception:
                    continue

                # Equipos ESPN → canon ES
                teams = {}
                for c in comp["competitors"]:
                    teams[c["homeAway"]] = espn_canon(c["team"]["displayName"])
                home_n    = teams.get("home", "")
                away_n    = teams.get("away", "")
                fecha_utc = ev["date"][:10]

                # Buscar fila en Fixture (fecha UTC y UTC-1)
                fechas = [fecha_utc]
                dt     = datetime.strptime(fecha_utc, "%Y-%m-%d")
                fechas.append((dt - timedelta(days=1)).strftime("%Y-%m-%d"))

                for f_try in fechas:
                    if (f_try, home_n, away_n) in fila_de_pend:
                        pid_found = fila_de_pend[(f_try, home_n, away_n)]
                        odds_result[str(pid_found)] = {
                            "pGanaA": pH, "pEmpate": pD, "pGanaB": pA,
                            "mlA": str(ml_home), "mlX": str(int(ml_draw)), "mlB": str(ml_away),
                        }
                        break
                    if (f_try, away_n, home_n) in fila_de_pend:
                        pid_found = fila_de_pend[(f_try, away_n, home_n)]
                        odds_result[str(pid_found)] = {
                            "pGanaA": pA, "pEmpate": pD, "pGanaB": pH,
                            "mlA": str(ml_away), "mlX": str(int(ml_draw)), "mlB": str(ml_home),
                        }
                        break

        except Exception as e:
            print(f"[AVISO] ESPN odds {fecha_str}: {e}")

        current += timedelta(days=1)

    odds_path = os.path.join(DATA, "odds.json")
    with open(odds_path, "w", encoding="utf-8") as fout:
        json.dump({
            "generado": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "fuente":   "DraftKings vía ESPN unofficial API",
            "partidos": odds_result,
        }, fout, ensure_ascii=False, indent=1)
    print(f"[OK] Cuotas guardadas: {len(odds_result)} partidos → Data/odds.json")

if __name__ == "__main__":
    main()
