# -*- coding: utf-8 -*-
"""
montecarlo.py — Simulación Monte Carlo del torneo restante (INSTRUCCIONES.md §8).
Correr SOLO cuando el usuario lo pida. Uso:
    python3 montecarlo.py [n_sims] [--dry]   (default 10000; --dry no escribe el xlsx)
Lee el estado actual del Tracker (Resultados, Equipos, Fixture), simula lo que falta
y escribe/reemplaza la hoja MonteCarlo con P(instancia) por selección, fecha y semilla.
"""
import os, re, sys, math, random
from datetime import datetime
sys.path.insert(0, os.path.dirname(__file__))
from comun import cargar_fixture
import numpy as np
from openpyxl import load_workbook

BASE = os.path.join(os.path.dirname(__file__), "..")
XLSX = os.path.join(BASE, "output", "Mundial2026_Tracker.xlsx")
SEMILLA = 20260612

def cargar_estado():
    wb = load_workbook(XLSX, data_only=False)
    wsE, wsR, wsF = wb["Equipos"], wb["Resultados"], wb["Fixture"]
    eq = {}
    for r in range(2, 50):
        t = wsE.cell(r, 1).value
        eq[t] = {"elo": float(wsE.cell(r, 3).value), "att": float(wsE.cell(r, 6).value),
                 "def": float(wsE.cell(r, 7).value)}
    mu = float(wsE.cell(2, 15).value); beta = float(wsE.cell(3, 15).value)
    K = float(wsE.cell(4, 15).value)
    res = {}
    for r in range(2, 106):
        ga, gb = wsR.cell(r, 2).value, wsR.cell(r, 3).value
        if ga is not None:
            res[wsR.cell(r, 1).value] = (int(ga), int(gb), wsR.cell(r, 4).value,
                                         wsR.cell(r, 5).value)
    fx = cargar_fixture(os.path.join(BASE, "Data", "worldcup.json"))
    return eq, mu, beta, K, res, fx

def lam(mu, beta, eq, a, b, loc_a, loc_b):
    dr = (eq[a]["elo"] + 100*loc_a) - (eq[b]["elo"] + 100*loc_b)
    d = beta * dr
    tot = min(3.6, max(1.8, mu * math.sqrt(eq[a]["att"]*eq[b]["att"]*eq[a]["def"]*eq[b]["def"])))
    return max(0.2, (tot+d)/2), max(0.2, (tot-d)/2)

def elo_update(eq, a, b, ga, gb, loc_a, loc_b, K):
    dr = (eq[a]["elo"] + 100*loc_a) - (eq[b]["elo"] + 100*loc_b)
    ea = 1/(1+10**(-dr/400))
    s = 1.0 if ga > gb else (0.5 if ga == gb else 0.0)
    d = abs(ga-gb)
    g = 1 if d <= 1 else (1.5 if d == 2 else (1.75 if d == 3 else 1.75+(d-3)/8))
    delta = K*g*(s-ea)
    eq[a]["elo"] += delta; eq[b]["elo"] -= delta

def orden_grupo(tabla, rng):
    # tabla: {eq: [pts, dg, gf]} -> orden FIFA aproximado (Pts, DG, GF, sorteo)
    return sorted(tabla, key=lambda t: (tabla[t][0], tabla[t][1], tabla[t][2], rng.random()),
                  reverse=True)

def main():
    n = 10000; dry = False
    for a in sys.argv[1:]:
        if a == "--dry": dry = True
        elif a.isdigit(): n = int(a)
    eq0, mu, beta, K, res, fx = cargar_estado()
    rng = random.Random(SEMILLA); nprng = np.random.default_rng(SEMILLA)
    grupos = {}
    for p in fx:
        if p["grupo"]:
            grupos.setdefault(p["grupo"], set()).update([p["a"], p["b"]])
    ko = sorted([p for p in fx if p["num"]], key=lambda p: p["num"])
    etapas = ["R32","Octavos","Cuartos","Semis","Final","Campeón"]
    cont = {t: {e: 0 for e in etapas} for t in eq0}

    for sim in range(n):
        eq = {t: dict(v) for t, v in eq0.items()}
        tabla = {g: {t: [0,0,0] for t in grupos[g]} for g in grupos}
        # fase de grupos
        for p in fx:
            if not p["grupo"]: continue
            a, b, g = p["a"], p["b"], p["grupo"]
            la = 1 if p["pais"] == a else 0; lb = 1 if p["pais"] == b else 0
            if p["id"] in res:
                ga, gb = res[p["id"]][0], res[p["id"]][1]
            else:
                lA, lB = lam(mu, beta, eq, a, b, la, lb)
                ga, gb = nprng.poisson(lA), nprng.poisson(lB)
            for t, gf, gc in ((a, ga, gb), (b, gb, ga)):
                tabla[g][t][0] += 3 if gf > gc else (1 if gf == gc else 0)
                tabla[g][t][1] += gf - gc; tabla[g][t][2] += gf
            elo_update(eq, a, b, ga, gb, la, lb, K)
        pos = {}; terceros = []
        for g in grupos:
            o = orden_grupo(tabla[g], rng)
            pos[f"1{g}"], pos[f"2{g}"] = o[0], o[1]
            terceros.append((g, o[2]))
        terceros.sort(key=lambda x: (tabla[x[0]][x[1]][0], tabla[x[0]][x[1]][1],
                                     tabla[x[0]][x[1]][2], rng.random()), reverse=True)
        clasif3 = dict(terceros[:8])      # grupo -> equipo
        for t in list(pos.values()) + list(clasif3.values()):
            cont[t]["R32"] += 1
        # bracket
        ganador, perdedor = {}, {}
        usados3 = set()
        for p in ko:
            equipos_m = []
            for code in (p["a"], p["b"]):
                if re.fullmatch(r"[12][A-L]", code):
                    equipos_m.append(pos[code])
                elif code.startswith("3"):
                    elegibles = [g for g in code[1:].split("/") if g in clasif3 and g not in usados3]
                    if not elegibles:
                        elegibles = [g for g in clasif3 if g not in usados3]
                    g = elegibles[0]; usados3.add(g)
                    equipos_m.append(clasif3[g])
                else:
                    m = re.fullmatch(r"([WL])(\d+)", code)
                    equipos_m.append((ganador if m.group(1) == "W" else perdedor)[int(m.group(2))])
            a, b = equipos_m
            la = 1 if p["pais"] == a else 0; lb = 1 if p["pais"] == b else 0
            if p["id"] in res:
                ga, gb, defin, gpen = res[p["id"]]
                w = a if ga > gb else (b if gb > ga else (gpen if gpen in (a, b) else rng.choice([a, b])))
            else:
                lA, lB = lam(mu, beta, eq, a, b, la, lb)
                ga, gb = nprng.poisson(lA), nprng.poisson(lB)
                w = a if ga > gb else (b if gb > ga else rng.choice([a, b]))  # penales ~ moneda
            elo_update(eq, a, b, ga, gb, la, lb, K)
            ganador[p["num"]], perdedor[p["num"]] = w, (b if w == a else a)
            siguiente = {"R32":"Octavos","Octavos":"Cuartos","Cuartos":"Semis",
                         "Semis":"Final"}.get(p["fase"])
            if siguiente:
                cont[w][siguiente] += 1
            elif p["fase"] == "Final":
                cont[w]["Campeón"] += 1

    filas = sorted(cont, key=lambda t: (-cont[t]["Campeón"], -cont[t]["Final"], -cont[t]["R32"]))
    print(f"{'Equipo':22}" + "".join(f"{e:>9}" for e in etapas))
    for t in filas[:15]:
        print(f"{t:22}" + "".join(f"{cont[t][e]/n:>9.1%}" for e in etapas))
    if dry:
        print("(--dry: no se escribió el xlsx)"); return
    wb = load_workbook(XLSX)
    if "MonteCarlo" in wb.sheetnames:
        del wb["MonteCarlo"]
    ws = wb.create_sheet("MonteCarlo")
    ws["A1"] = f"MONTE CARLO — {n} simulaciones | {datetime.now():%Y-%m-%d %H:%M} | semilla {SEMILLA} | P(pasa de grupo)=P(R32)"
    ws.append(["Equipo"] + [f"P({e})" for e in etapas])
    for t in filas:
        ws.append([t] + [cont[t][e]/n for e in etapas])
    for row in ws.iter_rows(min_row=3, min_col=2):
        for c in row:
            c.number_format = "0.0%"
    wb.save(XLSX)
    print("[OK] Hoja MonteCarlo escrita.")

if __name__ == "__main__":
    main()
