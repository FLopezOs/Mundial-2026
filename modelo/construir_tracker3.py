# -*- coding: utf-8 -*-
"""Parte 3: hoja Tablas (grupos en vivo, mejores terceros, bracket) + formato condicional."""
import os, re, sys
sys.path.insert(0, os.path.dirname(__file__))
from comun import cargar_fixture
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import FormulaRule

BASE = os.path.join(os.path.dirname(__file__), "..")
F_HDR = Font(name="Arial", size=10, bold=True, color="FFFFFF")
F_STD = Font(name="Arial", size=10)
F_TIT = Font(name="Arial", size=11, bold=True)
F_NOTA = Font(name="Arial", size=9, italic=True, color="808080")
FILL_HDR = PatternFill("solid", start_color="1F4E78")
CENTRO = Alignment(horizontal="center")

fx = cargar_fixture(os.path.join(BASE, "Data", "worldcup.json"))
wb = load_workbook(os.path.join(BASE, "output", "Mundial2026_Tracker.xlsx"))
ws = wb.create_sheet("Tablas")

# equipos por grupo (orden de aparición en el fixture)
grupos = {}
for p in fx:
    if p["grupo"]:
        for t in (p["a"], p["b"]):
            grupos.setdefault(p["grupo"], [])
            if t not in grupos[p["grupo"]]:
                grupos[p["grupo"]].append(t)
letras = sorted(grupos)

ws.cell(row=1, column=1, value="TABLAS DE GRUPOS (en vivo desde Resultados) — Desempate implementado: Pts → DG → GF. Head-to-head y fair play NO automatizados: verificar manualmente ante empate exacto.").font = F_NOTA

MR = "Modelo!$C$3:$C$106"; MD = "Modelo!$D$3:$D$106"; ME = "Modelo!$E$3:$E$106"
MH = "Modelo!$H$3:$H$106"; MI = "Modelo!$I$3:$I$106"; MJ = "Modelo!$J$3:$J$106"
G = f'({MR}="Grupos")'

filas_grupo = {}   # letra -> (fila_eq1, fila_eq4)
r = 3
for g in letras:
    ws.cell(row=r, column=1, value=f"GRUPO {g}").font = F_TIT
    for j, t in enumerate(["Equipo","PJ","PG","PE","PP","GF","GC","DG","Pts","Pos","Score"], 1):
        c = ws.cell(row=r+1, column=j, value=t)
        c.font, c.fill, c.alignment = F_HDR, FILL_HDR, CENTRO
    r0 = r + 2
    for k, t in enumerate(grupos[g]):
        rr = r0 + k
        ws.cell(row=rr, column=1, value=t).font = F_STD
        A = f"$A{rr}"
        ws.cell(row=rr, column=2, value=f"=SUMPRODUCT({G}*(({MD}={A})+({ME}={A}))*{MH})")
        ws.cell(row=rr, column=3, value=f"=SUMPRODUCT({G}*({MD}={A})*{MH}*({MI}>{MJ}))+SUMPRODUCT({G}*({ME}={A})*{MH}*({MJ}>{MI}))")
        ws.cell(row=rr, column=4, value=f"=SUMPRODUCT({G}*(({MD}={A})+({ME}={A}))*{MH}*({MI}={MJ}))")
        ws.cell(row=rr, column=5, value=f"=$B{rr}-$C{rr}-$D{rr}")
        ws.cell(row=rr, column=6, value=f"=SUMPRODUCT({G}*({MD}={A})*{MH}*{MI})+SUMPRODUCT({G}*({ME}={A})*{MH}*{MJ})")
        ws.cell(row=rr, column=7, value=f"=SUMPRODUCT({G}*({MD}={A})*{MH}*{MJ})+SUMPRODUCT({G}*({ME}={A})*{MH}*{MI})")
        ws.cell(row=rr, column=8, value=f"=$F{rr}-$G{rr}")
        ws.cell(row=rr, column=9, value=f"=3*$C{rr}+$D{rr}")
        ws.cell(row=rr, column=10, value=f"=1+SUMPRODUCT(($K${r0}:$K${r0+3}>$K{rr})*1)")
        ws.cell(row=rr, column=11, value=f"=$I{rr}*1000000+($H{rr}+200)*1000+$F{rr}")
        for j in range(2, 12):
            ws.cell(row=rr, column=j).font = F_STD
            ws.cell(row=rr, column=j).alignment = CENTRO
    filas_grupo[g] = (r0, r0 + 3)
    r += 7

# ---- mejores terceros ----
ws.cell(row=r, column=1, value="MEJORES TERCEROS (clasifican los 8 primeros) — Pts → DG → GF; fair play y ranking FIFA no automatizados.").font = F_TIT
for j, t in enumerate(["Grupo","Equipo","Pts","DG","GF","Score","Rank","Clasifica"], 1):
    c = ws.cell(row=r+1, column=j, value=t)
    c.font, c.fill, c.alignment = F_HDR, FILL_HDR, CENTRO
r0t = r + 2
for k, g in enumerate(letras):
    rr = r0t + k
    f1, f4 = filas_grupo[g]
    ws.cell(row=rr, column=1, value=g).font = F_STD
    ws.cell(row=rr, column=2, value=f"=IFERROR(INDEX($A${f1}:$A${f4},MATCH(3,$J${f1}:$J${f4},0)),\"\")")
    ws.cell(row=rr, column=3, value=f"=IFERROR(INDEX($I${f1}:$I${f4},MATCH(3,$J${f1}:$J${f4},0)),\"\")")
    ws.cell(row=rr, column=4, value=f"=IFERROR(INDEX($H${f1}:$H${f4},MATCH(3,$J${f1}:$J${f4},0)),\"\")")
    ws.cell(row=rr, column=5, value=f"=IFERROR(INDEX($F${f1}:$F${f4},MATCH(3,$J${f1}:$J${f4},0)),\"\")")
    ws.cell(row=rr, column=6, value=f"=IFERROR(INDEX($K${f1}:$K${f4},MATCH(3,$J${f1}:$J${f4},0)),\"\")")
    ws.cell(row=rr, column=7, value=f"=IF($F{rr}=\"\",\"\",1+SUMPRODUCT(($F${r0t}:$F${r0t+11}>$F{rr})*1))")
    ws.cell(row=rr, column=8, value=f"=IF($G{rr}=\"\",\"\",IF($G{rr}<=8,\"Sí\",\"No\"))")
    for j in range(2, 9):
        ws.cell(row=rr, column=j).font = F_STD
        ws.cell(row=rr, column=j).alignment = CENTRO
fila_terc = (r0t, r0t + 11)
r = r0t + 13

# ---- bracket ----
ws.cell(row=r, column=1, value="BRACKET ELIMINATORIO — Estado: REAL (resultado cargado) / PROYECCIÓN (modelo: avanza el mayor Elo vigente; en terceros, el mejor rankeado elegible — la asignación oficial FIFA puede diferir y puede duplicar equipos hasta que cierren los grupos).").font = F_TIT
for j, t in enumerate(["Num","ID","Ronda","Cod_A","Cod_B","Equipo_A","Equipo_B","Ganador","Perdedor","Estado"], 1):
    c = ws.cell(row=r+1, column=j, value=t)
    c.font, c.fill, c.alignment = F_HDR, FILL_HDR, CENTRO
ko = sorted([p for p in fx if p["num"]], key=lambda p: p["num"])
r0b = r + 2
fila_num = {p["num"]: r0b + i for i, p in enumerate(ko)}

def formula_codigo(code, rr, col_destino):
    """Fórmula que resuelve un código de bracket a un equipo."""
    if re.fullmatch(r"[12][A-L]", code):
        pos, g = int(code[0]), code[1]
        f1, f4 = filas_grupo[g]
        return f"=IFERROR(INDEX($A${f1}:$A${f4},MATCH({pos},$J${f1}:$J${f4},0)),\"\")"
    if code.startswith("3"):
        lets = code[1:].split("/")
        t1, t2 = fila_terc
        # ranks de los grupos elegibles en columnas auxiliares L..P de esta fila
        for x, le in enumerate(lets):
            ws.cell(row=rr, column=12+x, value=f"=IFERROR(VLOOKUP(\"{le}\",$A${t1}:$G${t2},7,0),99)").font = F_NOTA
        ws.cell(row=rr, column=17, value=f"=MIN(L{rr}:P{rr})").font = F_NOTA
        return f"=IFERROR(INDEX($B${t1}:$B${t2},MATCH($Q{rr},$G${t1}:$G${t2},0)),\"\")"
    m = re.fullmatch(r"([WL])(\d+)", code)
    if m:
        tipo, num = m.group(1), int(m.group(2))
        fr = fila_num[num]
        return f"=${'H' if tipo=='W' else 'I'}{fr}"
    raise ValueError(code)

for i, p in enumerate(ko):
    rr = r0b + i
    idr = p["id"] + 1  # fila en Resultados
    ws.cell(row=rr, column=1, value=p["num"])
    ws.cell(row=rr, column=2, value=p["id"])
    ws.cell(row=rr, column=3, value=p["fase"])
    ws.cell(row=rr, column=4, value=p["a"])
    ws.cell(row=rr, column=5, value=p["b"])
    ws.cell(row=rr, column=6, value=formula_codigo(p["a"], rr, 6))
    ws.cell(row=rr, column=7, value=formula_codigo(p["b"], rr, 7))
    gana_pen = f"IF(Resultados!$E{idr}<>\"\",Resultados!$E{idr},$F{rr})"
    real = f"IF(Resultados!$B{idr}>Resultados!$C{idr},$F{rr},IF(Resultados!$C{idr}>Resultados!$B{idr},$G{rr},{gana_pen}))"
    proy = f"IF(VLOOKUP($F{rr},Equipos!$A:$D,4,0)>=VLOOKUP($G{rr},Equipos!$A:$D,4,0),$F{rr},$G{rr})"
    ws.cell(row=rr, column=8, value=f"=IF(OR($F{rr}=\"\",$G{rr}=\"\"),\"\",IF(ISNUMBER(Resultados!$B{idr}),{real},{proy}))")
    real_p = f"IF(Resultados!$B{idr}>Resultados!$C{idr},$G{rr},IF(Resultados!$C{idr}>Resultados!$B{idr},$F{rr},IF(Resultados!$E{idr}<>\"\",IF(Resultados!$E{idr}=$F{rr},$G{rr},$F{rr}),$G{rr})))"
    proy_p = f"IF(VLOOKUP($F{rr},Equipos!$A:$D,4,0)>=VLOOKUP($G{rr},Equipos!$A:$D,4,0),$G{rr},$F{rr})"
    ws.cell(row=rr, column=9, value=f"=IF(OR($F{rr}=\"\",$G{rr}=\"\"),\"\",IF(ISNUMBER(Resultados!$B{idr}),{real_p},{proy_p}))")
    ws.cell(row=rr, column=10, value=f"=IF(OR($F{rr}=\"\",$G{rr}=\"\"),\"PENDIENTE\",IF(ISNUMBER(Resultados!$B{idr}),\"REAL\",\"PROYECCIÓN\"))")
    for j in range(1, 11):
        ws.cell(row=rr, column=j).font = F_STD
        ws.cell(row=rr, column=j).alignment = CENTRO

# formato condicional
verde = PatternFill("solid", start_color="C6EFCE")
ambar = PatternFill("solid", start_color="FFEB9C")
for g in letras:
    f1, f4 = filas_grupo[g]
    ws.conditional_formatting.add(f"A{f1}:J{f4}",
        FormulaRule(formula=[f"$J{f1}<=2"], fill=verde))
t1, t2 = fila_terc
ws.conditional_formatting.add(f"A{t1}:H{t2}", FormulaRule(formula=[f"$H{t1}=\"Sí\""], fill=verde))
rbN = r0b + len(ko) - 1
ws.conditional_formatting.add(f"A{r0b}:J{rbN}", FormulaRule(formula=[f"$J{r0b}=\"PROYECCIÓN\""], fill=ambar))
ws.conditional_formatting.add(f"A{r0b}:J{rbN}", FormulaRule(formula=[f"$J{r0b}=\"REAL\""], fill=verde))

for col, w in zip("ABCDEFGHIJ", [9,6,11,12,12,18,18,18,18,12]):
    ws.column_dimensions[col].width = w
for col in "LMNOPQ":
    ws.column_dimensions[col].width = 5

# orden de hojas y colores de pestaña
orden = ["Fixture","Resultados","Equipos","Modelo","Predicciones","Tablas","Matrices"]
wb._sheets = [wb[n] for n in orden]
for n, c in zip(orden, ["4472C4","FFC000","70AD47","A6A6A6","ED7D31","1F4E78","D9D9D9"]):
    wb[n].sheet_properties.tabColor = c

wb.save(os.path.join(BASE, "output", "Mundial2026_Tracker.xlsx"))
print("Parte 3 OK: Tablas + formato. Hojas:", wb.sheetnames)
