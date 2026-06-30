# -*- coding: utf-8 -*-
"""
exportar_json.py — Exporta el estado del Tracker a webapp/public/data.json (§11).
Correr DESPUÉS de cada actualización del Tracker (actualizar.py + recálculo del xlsx).
Lee valores calculados (data_only): el xlsx debe haberse recalculado (Excel o recalc.py).
Incluye: horaChile por partido, calibracion (μ, β, Att/Def) para recálculo dinámico.
Merge con Data/resultados_manuales.json para resultados que ESPN no capturó.
"""
import json, os, re, sys
from datetime import datetime
sys.path.insert(0, os.path.dirname(__file__))
from openpyxl import load_workbook
from comun import JSON_A_CANON
try:
    sys.stdout.reconfigure(encoding="utf-8")
except AttributeError:
    pass

BASE          = os.path.join(os.path.dirname(__file__), "..")
XLSX          = os.path.join(BASE, "output", "Mundial2026_Tracker.xlsx")
DEST          = os.path.join(BASE, "webapp", "public", "data.json")
WC_JSON       = os.path.join(BASE, "Data", "worldcup.json")
CALIB_JSON    = os.path.join(BASE, "Data", "calibracion.json")
ODDS_JSON     = os.path.join(BASE, "Data", "odds.json")
MANUALES_JSON = os.path.join(BASE, "Data", "resultados_manuales.json")
CANALES_JSON  = os.path.join(BASE, "Data", "canales.json")
BRACKET_JSON  = os.path.join(BASE, "Data", "bracket_manuales.json")

def num(v, nd=4):
    return round(float(v), nd) if isinstance(v, (int, float)) else None

def to_chile(hora_local, offset_utc):
    """Convierte hora local (HH:MM) con offset UTC a hora Chile (UTC-3)."""
    h, m = map(int, hora_local.split(':'))
    chile_h = (h - offset_utc - 4) % 24
    return f"{chile_h:02d}:{m:02d}"

def build_hora_map(wc_path):
    """Construye {(fecha, frozenset([t1,t2])): horaChile} desde worldcup.json."""
    hora_map = {}
    try:
        with open(wc_path, encoding="utf-8") as f:
            wc = json.load(f)
    except Exception:
        return hora_map
    for m in wc.get("matches", []):
        t_raw = m.get("time", "")
        fecha = str(m.get("date", ""))
        rx = re.match(r'(\d+:\d+)\s+UTC([+-]\d+)', t_raw)
        if not rx or not fecha:
            continue
        hora_chile = to_chile(rx.group(1), int(rx.group(2)))
        t1 = JSON_A_CANON.get(m.get("team1",""), m.get("team1",""))
        t2 = JSON_A_CANON.get(m.get("team2",""), m.get("team2",""))
        key = (fecha, frozenset([t1.lower(), t2.lower()]))
        hora_map[key] = hora_chile
    return hora_map

def build_odds_map(odds_path):
    """Carga Data/odds.json → dict {str(id): {pGanaA, pEmpate, pGanaB, mlA, mlX, mlB}}."""
    try:
        with open(odds_path, encoding="utf-8") as f:
            d = json.load(f)
        return d.get("partidos", {})
    except Exception:
        return {}

def build_canales_map(canales_path):
    """Carga Data/canales.json → dict {str(id): [canal, ...]}."""
    try:
        with open(canales_path, encoding="utf-8") as f:
            d = json.load(f)
        return d.get("partidos", {})
    except Exception:
        return {}

def build_bracket_map(bracket_path):
    """Carga Data/bracket_manuales.json → dict {str(id): {equipoA, equipoB}}."""
    try:
        with open(bracket_path, encoding="utf-8") as f:
            d = json.load(f)
        return d.get("partidos", {})
    except Exception:
        return {}

def build_manuales_map(manuales_path):
    """
    Carga Data/resultados_manuales.json → dict {str(id): {golesA, golesB, ...}}.
    Se aplica cuando el Excel tiene null para ese partido (ESPN no lo capturó).
    """
    try:
        with open(manuales_path, encoding="utf-8") as f:
            d = json.load(f)
        return d.get("partidos", {})
    except Exception:
        return {}

def build_calibracion(calib_path):
    """Carga calibracion.json y traduce nombres de equipos a canónicos ES."""
    try:
        with open(calib_path, encoding="utf-8") as f:
            c = json.load(f)
    except Exception:
        return None
    # Mapa inverso: nombre inglés → canon ES
    ing_a_canon = {k: v for k, v in JSON_A_CANON.items()}
    equipos_es = {}
    for nombre_en, vals in c.get("equipos", {}).items():
        nombre_es = ing_a_canon.get(nombre_en, nombre_en)
        equipos_es[nombre_es] = {
            "att": round(vals["att"], 3),
            "def": round(vals["def"], 3),
        }
    return {
        "mu":    round(c["mu"], 4),
        "beta":  round(c["beta"], 6),
        "equipos": equipos_es,
    }

def main():
    wb = load_workbook(XLSX, data_only=True)
    wsF, wsR, wsE, wsP = wb["Fixture"], wb["Resultados"], wb["Equipos"], wb["Predicciones"]

    hora_map     = build_hora_map(WC_JSON)
    calibracion  = build_calibracion(CALIB_JSON)
    odds_map     = build_odds_map(ODDS_JSON)
    manuales_map = build_manuales_map(MANUALES_JSON)
    canales_map  = build_canales_map(CANALES_JSON)
    bracket_map  = build_bracket_map(BRACKET_JSON)

    if manuales_map:
        print(f"[OK] resultados_manuales.json: {len(manuales_map)} partido(s) cargado(s) → "
              f"{sorted(int(k) for k in manuales_map)}")

    equipos = []
    for r in range(2, 50):
        equipos.append({
            "nombre":        wsE.cell(r, 1).value,
            "confederacion": wsE.cell(r, 2).value,
            "eloInicial":    num(wsE.cell(r, 3).value, 1),
            "eloActual":     num(wsE.cell(r, 4).value, 1),
        })

    partidos = []
    manuales_aplicados = 0
    for r in range(2, 106):
        pid   = wsF.cell(r, 1).value
        fecha = wsF.cell(r, 2).value
        ea    = wsF.cell(r, 5).value or ""
        eb    = wsF.cell(r, 6).value or ""
        ea_orig, eb_orig = ea, eb
        if str(pid) in bracket_map:
            ea = bracket_map[str(pid)].get("equipoA", ea)
            eb = bracket_map[str(pid)].get("equipoB", eb)
        p = {
            "id":      pid,
            "numFifa": wsF.cell(r, 12).value,
            "fecha":   fecha,
            "fase":    wsF.cell(r, 3).value,
            "grupo":   wsF.cell(r, 4).value or None,
            "equipoA": ea,
            "equipoB": eb,
            "ciudad":  wsF.cell(r, 8).value,
            "pais":    wsF.cell(r, 9).value,
        }
        # Hora Chile — intenta primero con nombres reales, luego con códigos originales del Excel
        key = (str(fecha), frozenset([ea.lower(), eb.lower()]))
        if key not in hora_map:
            key = (str(fecha), frozenset([ea_orig.lower(), eb_orig.lower()]))
        if key in hora_map:
            p["horaChile"] = hora_map[key]

        ga, gb = wsR.cell(r, 2).value, wsR.cell(r, 3).value

        if ga is not None and gb is not None:
            p["resultado"] = {
                "golesA":        int(ga),
                "golesB":        int(gb),
                "definidoPor":   wsR.cell(r, 4).value or "90",
                "ganadorPenales":wsR.cell(r, 5).value,
                "penScoreA":     wsR.cell(r, 6).value,
                "penScoreB":     wsR.cell(r, 7).value,
            }
        elif str(pid) in manuales_map:
            m = manuales_map[str(pid)]
            p["resultado"] = {
                "golesA":        m["golesA"],
                "golesB":        m["golesB"],
                "definidoPor":   m.get("definidoPor", "90"),
                "ganadorPenales":m.get("ganadorPenales"),
                "penScoreA":     m.get("penScoreA"),
                "penScoreB":     m.get("penScoreB"),
            }
            manuales_aplicados += 1

        rp = pid + 2
        pa = wsP.cell(rp, 6).value
        if isinstance(pa, (int, float)):
            # Probabilidades desde el Excel (requiere recalc.py o Excel abierto)
            p["modelo"] = {
                "pGanaA": num(pa),
                "pEmpate": num(wsP.cell(rp, 7).value),
                "pGanaB":  num(wsP.cell(rp, 8).value),
            }

        # Cuotas de casas de apuesta (solo partidos sin resultado)
        if "resultado" not in p and str(pid) in odds_map:
            p["oddsImplied"] = odds_map[str(pid)]

        # Canales de transmisión en Chile
        if str(pid) in canales_map:
            p["canales"] = canales_map[str(pid)]

        partidos.append(p)

    data = {
        "generado":    datetime.now().strftime("%Y-%m-%d %H:%M"),
        "torneo":      "Mundial 2026",
        "nota":        "Generado por modelo/exportar_json.py — no editar a mano.",
        "calibracion": calibracion,
        "equipos":     equipos,
        "partidos":    partidos,
    }
    os.makedirs(os.path.dirname(DEST), exist_ok=True)
    with open(DEST, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=1)
    jugados  = sum(1 for p in partidos if "resultado" in p)
    con_hora = sum(1 for p in partidos if "horaChile" in p)
    con_odds = sum(1 for p in partidos if "oddsImplied" in p)
    if manuales_aplicados:
        print(f"[OK] Resultados manuales aplicados: {manuales_aplicados} partido(s).")
    print(f"[OK] {len(partidos)} partidos ({jugados} jugados), {con_hora} con hora Chile, "
          f"{con_odds} con cuotas, calibración={'sí' if calibracion else 'NO'}.")

if __name__ == "__main__":
    main()
