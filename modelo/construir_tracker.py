# -*- coding: utf-8 -*-
"""
construir_tracker.py — Primera construcción de output/Mundial2026_Tracker.xlsx (§4).
Fórmulas vivas por partido; valores calculados (Elo inicial, Att/Def, μ, β) pegados
solo en Equipos con fecha anotada (excepción permitida §4).
"""
import json, os, sys
sys.path.insert(0, os.path.dirname(__file__))
from comun import CONFED, CANON_A_RESULTS, CANON_A_JSON, cargar_fixture
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

BASE = os.path.join(os.path.dirname(__file__), "..")
DATA = os.path.join(BASE, "Data")
OUT = os.path.join(BASE, "output")
FECHA_CALC = "2026-06-12"

# ---------- estilos ----------
F_HDR = Font(name="Arial", size=10, bold=True, color="FFFFFF")
F_STD = Font(name="Arial", size=10)
F_TIT = Font(name="Arial", size=11, bold=True)
F_NOTA = Font(name="Arial", size=9, italic=True, color="808080")
FILL_HDR = PatternFill("solid", start_color="1F4E78")
FILL_IN = PatternFill("solid", start_color="FFF2CC")
THIN = Border(*[Side(style="thin", color="D9D9D9")]*4)
CENTRO = Alignment(horizontal="center")

def hdr(ws, row, cols, start=1):
    for j, t in enumerate(cols, start):
        c = ws.cell(row=row, column=j, value=t)
        c.font, c.fill, c.alignment = F_HDR, FILL_HDR, CENTRO

def estilo_rango(ws, r1, r2, c1, c2):
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            cell = ws.cell(row=r, column=c)
            if cell.font.name != "Arial":
                cell.font = F_STD
            cell.border = THIN

# ---------- datos ----------
fx = cargar_fixture(os.path.join(DATA, "worldcup.json"))
cal = json.load(open(os.path.join(DATA, "calibracion.json"), encoding="utf-8"))
import csv
elo0 = {}
with open(os.path.join(DATA, "elo_calculado.csv"), encoding="utf-8") as f:
    for row in csv.DictReader(f):
        elo0[row["equipo"]] = float(row["elo"])

equipos = sorted(CONFED.keys(), key=lambda t: -elo0[CANON_A_RESULTS[t]])
NEQ = len(equipos)
assert NEQ == 48

wb = Workbook()
wb.remove(wb.active)

# ============ Hoja Fixture ============
ws = wb.create_sheet("Fixture")
hdr(ws, 1, ["ID_Partido","Fecha","Fase","Grupo","Equipo_A","Equipo_B","Estadio","Ciudad","Pais_Sede","Localia_A","Localia_B","Num_FIFA"])
for p in fx:
    r = p["id"] + 1
    ws.cell(row=r, column=1, value=p["id"])
    ws.cell(row=r, column=2, value=p["fecha"])
    ws.cell(row=r, column=3, value=p["fase"])
    ws.cell(row=r, column=4, value=p["grupo"])
    ws.cell(row=r, column=5, value=p["a"])
    ws.cell(row=r, column=6, value=p["b"])
    ws.cell(row=r, column=8, value=p["ciudad"])
    ws.cell(row=r, column=9, value=p["pais"])
    for col, eq in ((10, "E"), (11, "F")):
        ws.cell(row=r, column=col, value=f'=IF(OR(AND(${eq}{r}="México",$I{r}="México"),AND(${eq}{r}="Estados Unidos",$I{r}="Estados Unidos"),AND(${eq}{r}="Canadá",$I{r}="Canadá")),1,0)')
    if p["num"]:
        ws.cell(row=r, column=12, value=p["num"])
estilo_rango(ws, 2, 105, 1, 12)
for col, w in zip("ABCDEFGHIJKL", [10,11,9,7,20,20,12,22,14,9,9,9]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A2"

# ============ Hoja Resultados ============
ws = wb.create_sheet("Resultados")
hdr(ws, 1, ["ID_Partido","Goles_A","Goles_B","Definido_Por","Ganador_Penales","Equipo_A (ref)","Equipo_B (ref)"])
for p in fx:
    r = p["id"] + 1
    ws.cell(row=r, column=1, value=p["id"])
    for c in (2, 3, 4, 5):
        ws.cell(row=r, column=c).fill = FILL_IN
    ws.cell(row=r, column=6, value=f"=VLOOKUP($A{r},Fixture!$A:$F,5,0)")
    ws.cell(row=r, column=7, value=f"=VLOOKUP($A{r},Fixture!$A:$F,6,0)")
estilo_rango(ws, 2, 105, 1, 7)
ws.cell(row=107, column=1, value="Única hoja de entrada manual. Goles_A/B: marcador final incl. alargue, SIN penales. Definido_Por: 90/ET/PEN. Si PEN, registrar Ganador_Penales con el nombre canónico.").font = F_NOTA
for col, w in zip("ABCDEFG", [10,9,9,12,16,20,20]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A2"

# ============ Hoja Equipos ============
ws = wb.create_sheet("Equipos")
hdr(ws, 1, ["Equipo","Confederacion","Elo_Inicial","Elo_Actual","PJ","Att","Def","Lambda_Base"])
for i, t in enumerate(equipos):
    r = i + 2
    en = CANON_A_RESULTS[t]
    eq = cal["equipos"][en]
    ws.cell(row=r, column=1, value=t)
    ws.cell(row=r, column=2, value=CONFED[t])
    ws.cell(row=r, column=3, value=round(elo0[en], 1))
    ws.cell(row=r, column=4, value=f"=INDEX(Modelo!$AE$106:$BZ$106,MATCH($A{r},Modelo!$AE$1:$BZ$1,0))")
    ws.cell(row=r, column=5, value=f"=SUMPRODUCT((Modelo!$D$3:$D$106=$A{r})*Modelo!$H$3:$H$106)+SUMPRODUCT((Modelo!$E$3:$E$106=$A{r})*Modelo!$H$3:$H$106)")
    ws.cell(row=r, column=6, value=eq["att"])
    ws.cell(row=r, column=7, value=eq["def"])
    ws.cell(row=r, column=8, value=eq["lambda_base"])
    ws.cell(row=r, column=3).number_format = "0.0"
    ws.cell(row=r, column=4).number_format = "0.0"
estilo_rango(ws, 2, 49, 1, 8)
# tabla de equivalencias
ws.cell(row=1, column=10, value="EQUIVALENCIAS DE NOMBRES").font = F_TIT
hdr(ws, 2, ["Canonico","results.csv","worldcup.json"], start=10)
for i, t in enumerate(equipos):
    r = i + 3
    ws.cell(row=r, column=10, value=t)
    ws.cell(row=r, column=11, value=CANON_A_RESULTS[t])
    ws.cell(row=r, column=12, value=CANON_A_JSON[t])
estilo_rango(ws, 3, 50, 10, 12)
# parámetros del modelo
ws.cell(row=1, column=14, value="PARÁMETROS DEL MODELO").font = F_TIT
params = [("mu (goles/partido)", cal["mu"]), ("beta (goles por punto Elo)", cal["beta"]),
          ("K Mundial", 60), ("Fecha de cálculo", FECHA_CALC),
          ("Ventana calibración", cal["ventana"])]
for i, (k, v) in enumerate(params):
    ws.cell(row=2+i, column=14, value=k).font = F_STD
    c = ws.cell(row=2+i, column=15, value=v)
    c.font = Font(name="Arial", size=10, color="0000FF")
ws.cell(row=8, column=14, value=f"Elo_Inicial, Att, Def y Lambda_Base calculados el {FECHA_CALC} por modelo/elo_historico.py y modelo/calibrar.py (results.csv 1872→2026-06-10). Única excepción a la regla de no pegar valores (§4). Fuente Elo: calculado propio (decisión del usuario 2026-06-12; xlsx aportado quedó de respaldo en Data/).").font = F_NOTA
for col, w in zip(["A","B","C","D","E","F","G","H","J","K","L","N","O"], [20,14,11,11,5,7,7,12,20,22,22,26,14]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A2"

wb.save(os.path.join(OUT, "Mundial2026_Tracker.xlsx"))
print("Parte 1 OK: Fixture, Resultados, Equipos")
